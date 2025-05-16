import tkinter as tk
from tkinter import ttk
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from tkinter import messagebox

import time
import requests
import os
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook, load_workbook
import threading


load_dotenv()

def post_to_webhook(data, webhook_url="http://localhost:5000/webhook"):
    """
    Envía un JSON con los datos de sensores al webhook.
    """
    try:
        resp = requests.post(webhook_url, json=data, timeout=5)
        resp.raise_for_status()
        print(f"[Webhook] Enviado: {data}")
    except requests.RequestException as e:
        print(f"[Webhook] Error enviando datos: {e}")

def send_text(bot_message):  # Envía mensajes al chatbot de Telegram
    bot_token = os.getenv('TELEGRAM_BOT_TOKEN')
    chat_ID = os.getenv('TELEGRAM_CHAT_ID')
    send_text_url = f'https://api.telegram.org/bot{bot_token}/sendMessage?chat_id={chat_ID}&parse_mode=Markdown&text={bot_message}'
    res = requests.post(send_text_url)
    return res
def send_image(image):                                            #funcion que envia imagenes al chatbot de telegram
    bot_token = os.getenv('TELEGRAM_BOT_TOKEN')
    chat_ID = os.getenv('TELEGRAM_CHAT_ID')
    send_image='https://api.telegram.org/bot' + bot_token + '/sendPhoto'
    data={'chat_id':chat_ID}
    files={'photo':(image,open(image,'rb'))}
    response=requests.post(send_image, files=files, data=data, verify=False)
    return response

class Scraping:
    def __init__(self):
        self.url = ''
        self.running = False
        self.options = Options()
        # Si deseas ejecutar sin ventana, descomenta las siguientes líneas:
        # self.options.add_argument("headless")
        self.options.add_argument("disable-gpu")
        self.options.add_argument("no-sandbox")
        self.id = os.getenv('accountId')
        self.user = os.getenv('userId')
        self.password = os.getenv('password')
        self.driver = webdriver.Chrome(options=self.options)
    

    def access(self, url):
        try:
            self.url = url
            self.driver.get(self.url)
        except Exception as e:
            print(f'Error obteniendo el driver: {e}')
            send_text('Error obteniendo el driver')
            self.driver.quit()
            return False

        try:

            id_cuenta = self.driver.find_element(by=By.ID, value='AccountId')
            user = self.driver.find_element(by=By.ID, value='UserId')
            password = self.driver.find_element(by=By.ID, value='Password')
            submit = self.driver.find_element(by=By.ID, value='submitBtn')


        

            id_cuenta.send_keys(self.id)
            user.send_keys(self.user)
            password.send_keys(self.password)
            submit.click()

            time.sleep(2)
            return True
        except Exception as e:
            print(f'No fue posible acceder: {e}')
            send_text('No fue posible acceder')
            self.driver.quit()
            return False

    def get_data(self, fecha, hora):
        data = self.driver.find_elements(by=By.CLASS_NAME, value='value')

        # Verifica si se han encontrado suficientes datos
        if not data:
            send_text("no se encontraron suficientes datos en la pagina")
            raise ValueError("No se encontraron suficientes datos en la página") 
            
            
        try:
            corr1 = float(data[0].text or 0)
            corr2 = float(data[1].text or 0)
            corr3 = float(data[2].text or 0)
        except Exception as e:
            print(f'Error al convertir datos: {e}')
            send_text(f'Error al convertir datos: {e}')
            return None

        prom = (corr1 + corr2 + corr3) / 3

        try:
            title = app.file_entry.get().upper()
            file_name = title + '.xlsx'
            if title=='':
                messagebox.showerror('Error',"INGRESE EL NOMBRE DEL ARCHIVO")
                app.stop_test()

            if os.path.exists(file_name):
                book = load_workbook(f'Q:\PUBLIC\CO_MDE_PRUEBAS_PR\CALENTAMIENTO CON PYTHON\CORRIENTES_CALENTAMIENTO\{file_name}')
                sheet = book.active
                row = sheet.max_row + 1
            else:
                book = Workbook()
                sheet = book.active
                sheet['A1'] = 'FECHA'
                sheet['B1'] = 'HORA'
                sheet['C1'] = 'FASE U'
                sheet['D1'] = 'FASE V'
                sheet['E1'] = 'FASE W'
                sheet['F1'] = 'PROMEDIO'
                row = 2

            sheet[f'A{row}'] = fecha
            sheet[f'B{row}'] = hora
            sheet[f'C{row}'] = corr1
            sheet[f'D{row}'] = corr2
            sheet[f'E{row}'] = corr3
            sheet[f'F{row}'] = round(prom, 2)
            book.save(f'Q:\PUBLIC\CO_MDE_PRUEBAS_PR\CALENTAMIENTO CON PYTHON\CORRIENTES_CALENTAMIENTO/{file_name}')
            row += 1

            print(f'FECHA: {fecha}\nHORA: {hora}\nFase U: {corr1} A')
            print(f'FECHA: {fecha}\nHORA: {hora}\nFase V: {corr2} A')
            print(f'FECHA: {fecha}\nHORA: {hora}\nFase W: {corr3} A')
            print(f'El promedio de las corrientes es: {round(prom, 2)} A')

            send_text(f'FECHA: {fecha} \nHORA: {hora} \nFase U: {corr1} A \nFase V: {corr2} A \nFase W: {corr3} A \nPromedio: {round(prom, 2)}A')
            
            return {"fecha": fecha, "hora": hora, "corr1": corr1, "corr2": corr2, "corr3": corr3, "prom": round(prom, 2)}
        except Exception as e:
            print(f'Error al guardar datos: {e}')
            send_text(f'Error al guardar datos, se ha desconectado la plataforma: {e}')
            return None
    def save_screenshot(self):
        self.driver.save_screenshot('Q:\PUBLIC\CO_MDE_PRUEBAS_PR\CALENTAMIENTO CON PYTHON\IMAGENES_CALENTAMIENTO/captura_corrientes.png')
        send_image('Q:\PUBLIC\CO_MDE_PRUEBAS_PR\CALENTAMIENTO CON PYTHON\IMAGENES_CALENTAMIENTO/captura_corrientes.png')
class CurrentTestApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pruebas de Corriente")
        self.scraping = Scraping()

        # Frame superior (Título)
        self.top_frame = tk.Frame(self.root, height=50)
        self.top_frame.pack(side=tk.TOP, fill=tk.X)
        self.title_label = tk.Label(self.top_frame, text="Pruebas de Corriente", font=("Arial", 16))
        self.title_label.pack(pady=10)
        self.frame2 = tk.Frame(self.root)
        
        # Frame izquierdo (Botones)
        self.left_frame = tk.Frame(self.root, width=150)
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)
        self.frame2.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=50, pady=10)

        self.start_button = tk.Button(self.left_frame, text="Iniciar", command=self.start_test)
        self.start_button.pack(pady=5)

        self.stop_button = tk.Button(self.left_frame, text="Detener", command=self.stop_test)
        self.stop_button.pack(pady=5)

        self.file_label = tk.Label(self.frame2, text="Ingrese el nombre del archivo:", font=("Arial", 14))
        self.file_label.pack(anchor='w', padx=10, pady=(10, 0))

        self.file_entry = tk.Entry(self.frame2, width=35)
        self.file_entry.pack(anchor='w', padx=10, pady=(0, 10))

        # Frame derecho (Treeview para mostrar información)
        self.right_frame = tk.Frame(self.root)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.treeview = ttk.Treeview(self.right_frame, columns=("Fecha", "Hora", "Fase U", "Fase V", "Fase W", "Promedio"), show='headings')
        self.treeview.pack(fill=tk.BOTH, expand=True)

        # Definir encabezados de columnas
        self.treeview.heading("Fecha", text="Fecha")
        self.treeview.heading("Hora", text="Hora")
        self.treeview.heading("Fase U", text="Fase U (A)")
        self.treeview.heading("Fase V", text="Fase V (A)")
        self.treeview.heading("Fase W", text="Fase W (A)")
        self.treeview.heading("Promedio", text="Promedio")

    def start_test(self):
        print("Iniciando prueba de corriente...")
        self.scraping.running = True
        threading.Thread(target=self.run_scraping, daemon=True).start()
        # threading.Thread(target=self.keep_screen_on, daemon=True).start()

    def stop_test(self):
        print("Deteniendo prueba de corriente...")
        self.scraping.running = False
        self.scraping.driver.quit()

    def run_scraping(self):
        try:
            date1 = datetime.now()
            fecha1 = date1.date()
            hora1 = date1.strftime('%H:%M')
            url2 = 'https://cloud.gennect.net/app/Dashboard/?culture=es'
            if not self.scraping.access(url2):
                return
            time.sleep(10)
            data = self.scraping.get_data(fecha1, hora1)
            if data:
                self.update_treeview(data)
                #post_to_webhook(data)
            self.scraping.save_screenshot()
            while self.scraping.running:
                date = datetime.now()
                fecha = date.date()
                hora = date.strftime('%H:%M')
                minute = date.minute
                second = date.second
                if minute in [0, 10, 20, 30, 40, 50] and second == 0:
                    data = self.scraping.get_data(fecha, hora)
                    if data:
                        self.update_treeview(data)
                    else:
                        print("No hay data")
                        send_text("No hay data")
                    self.scraping.save_screenshot()
                time.sleep(1)  # Pequeña pausa para evitar bucle muy acelerado
        except Exception as e:
            print(f'Error durante el scraping: {e}')
            send_text(f'Error durante el scraping: {e}')
        finally:
            self.scraping.driver.quit()  # Aseguramos que se libere la sesión

                
        

    def update_treeview(self, data):
        self.treeview.insert("", "end", values=(data["fecha"], data["hora"], data["corr1"], data["corr2"], data["corr3"], data["prom"]))

if __name__ == "__main__":
    root = tk.Tk()
    app = CurrentTestApp(root)
    root.mainloop()